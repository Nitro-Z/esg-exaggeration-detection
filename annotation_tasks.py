import os
import re
import random
import pandas as pd
import nltk
nltk.download('punkt')
nltk.download('punkt_tab')
from nltk.tokenize import sent_tokenize

annotation_dir = r'C:\Users\Administrator\Desktop\ESG_Project\Dataset\annotation_sample'
output_file    = r'C:\Users\Administrator\Desktop\ESG_Project\Dataset\annotation_task.xlsx'

N_SAMPLE    = 300
SEED        = 42
MIN_WORDS   = 8 
MAX_WORDS   = 60

def clean_for_excel(text):
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', ' ', text)
    text = re.sub(r'[\ufffe\uffff\ufdd0-\ufdef]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

print("Đang đọc và tokenize...")
all_sentences = []

file_list = [f for f in os.listdir(annotation_dir) if f.lower().endswith('.txt')]

for filename in file_list:
    fpath = os.path.join(annotation_dir, filename)
    try:
        with open(fpath, 'r', encoding='utf-8') as f:
            text = f.read()
    except Exception as e:
        print(f"  ⚠️ Bỏ qua {filename}: {e}")
        continue

    for sent in sent_tokenize(text):
        sent = clean_for_excel(sent.strip())
        words = sent.split()

        if len(words) < MIN_WORDS:                     continue
        if len(words) > MAX_WORDS:                     continue
        if '....' in sent:                             continue
        if sent.isupper():                             continue
        if re.match(r'^[\d\s\.\-\|\•\*\/]+$', sent):  continue
        if sent.count('@') > 0:                        continue
        if sent.count('•') > 2:                        continue

        all_sentences.append({'filename': filename, 'sentence': sent})

print(f"Tổng câu hợp lệ: {len(all_sentences):,}")

random.seed(SEED)
sampled = random.sample(all_sentences, min(N_SAMPLE, len(all_sentences)))

df = pd.DataFrame(sampled)
df.insert(0, 'id', range(1, len(df) + 1))
df['label'] = ''
df['note']  = ''
df = df[['id', 'filename', 'sentence', 'label', 'note']]
df['sentence'] = df['sentence'].apply(clean_for_excel)
df['filename'] = df['filename'].apply(clean_for_excel)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Annotation')
    ws = writer.sheets['Annotation']

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    label_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.fill = label_fill
            cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'

    guide = writer.book.create_sheet('Labeling Guide')
    instructions = [
        ['HƯỚNG DẪN LABEL'],
        [''],
        ['Cột label: điền 0 hoặc 1'],
        [''],
        ['label = 1 (EXAGGERATION) khi câu có:'],
        ['  - Tuyên bố tuyệt đối không verify được: "world-class", "best in class", "zero emissions"'],
        ['  - Superlative không có bằng chứng: "unprecedented", "unparalleled", "global leader"'],
        ['  - Cam kết mơ hồ không có target/deadline: "committed to sustainability", "we strive to"'],
        ['  - Phóng đại rõ ràng: "revolutionary", "groundbreaking", "paradigm-shifting"'],
        ['  - Tự khen không căn cứ: "exemplary", "industry-leading", "gold standard"'],
        [''],
        ['label = 0 (NORMAL) khi câu:'],
        ['  - Có số liệu cụ thể: "reduced emissions by 23% in 2023"'],
        ['  - Mô tả hành động cụ thể: "installed 500 solar panels at headquarters"'],
        ['  - Có hedge/qualifier: "targeting net zero by 2050", "aiming to reduce by 30%"'],
        ['  - Ngôn ngữ trung tính, factual'],
        [''],
        ['Nếu không chắc → label = 0 (conservative)'],
        ['Cột note: ghi lý do nếu muốn, không bắt buộc'],
    ]
    for row in instructions:
        guide.append(row)
    guide.column_dimensions['A'].width = 80

print(f"\n✅ Xuất xong: {output_file}")
print(f"   {len(df)} câu cần label")
print(f"\n📋 Mở file → điền 0 hoặc 1 vào cột D (màu vàng)")
print(f"   ~10 giây/câu → 300 câu ≈ 50 phút")